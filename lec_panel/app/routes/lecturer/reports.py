"""
Enhanced Reports Routes - With Semester/Year Filtering and Improved UX
Supports semester-based filtering (Sem 1: Sep-Dec, Sem 2: Jan-Apr)
"""

from flask import Blueprint, render_template, request, jsonify, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from app.services.report_service import ReportService
from app.utils.pdf_generator import PDFGenerator
from app.utils.excel_generator import ExcelGenerator
from app.decorators.auth import active_account_required, owns_session, owns_class
from datetime import datetime, timedelta
from collections import defaultdict
import io
import csv
import json

reports_bp = Blueprint('reports', __name__, url_prefix='/lecturer/reports')

# Initialize services
report_service = ReportService()
pdf_generator = PDFGenerator()
excel_generator = ExcelGenerator()


def get_current_semester():
    """
    Determine current semester based on date
    Semester 1: September - December
    Semester 2: January - April
    Holiday: May - August (defaults to Semester 2 for filtering)
    
    Returns:
        str: Current semester ('1' or '2')
    """
    now = datetime.now()
    month = now.month
    
    if month in [9, 10, 11, 12]:
        return '1'  # Semester 1
    elif month in [1, 2, 3, 4]:
        return '2'  # Semester 2
    else:
        # May-August is holiday period, default to last active semester
        return '2'


def organize_classes_by_semester(classes):
    """
    Organize classes by semester (all years in same semester grouped together)
    Semester format: year.semester (e.g., 3.1 = Year 3, Semester 1)
    
    Args:
        classes: List of Class objects
        
    Returns:
        tuple: (organized_dict, current_semester)
    """
    current_semester = get_current_semester()
    
    # Group classes by semester, then by year level
    organized = defaultdict(lambda: defaultdict(list))
    
    for cls in classes:
        # Extract semester from format: year.semester (e.g., '3.1' -> semester '1')
        if cls.semester and '.' in cls.semester:
            parts = cls.semester.split('.')
            semester = parts[1] if len(parts) > 1 else current_semester
            # Only include valid semesters (1 or 2)
            if semester not in ['1', '2']:
                continue
        elif cls.semester:
            # If no dot, assume it's just the semester number
            semester = cls.semester
            if semester not in ['1', '2']:
                continue
        else:
            semester = current_semester
        
        year = cls.year or 1
        organized[semester][year].append(cls)
    
    # Sort structure - only include semesters 1 and 2
    sorted_structure = {}
    for semester in ['1', '2']:  # Explicitly only include valid semesters
        if semester in organized:
            sorted_structure[semester] = {}
            for year in sorted(organized[semester].keys()):
                sorted_structure[semester][year] = sorted(
                    organized[semester][year],
                    key=lambda x: x.class_name
                )
    
    return sorted_structure, current_semester


def serialize_class(cls):
    """
    Convert a Class object to a JSON-serializable dictionary
    
    Args:
        cls: Class model object
        
    Returns:
        dict: Serialized class data
    """
    return {
        'class_id': cls.class_id,
        'class_name': cls.class_name,
        'course_code': cls.course_code,
        'year': cls.year,
        'semester': cls.semester
    }


@reports_bp.route('/')
@login_required
@active_account_required
def index():
    """
    Reports dashboard - Overview of available reports with smart filtering
    """
    # Get instructor's classes
    all_classes = current_user.get_assigned_classes()
    
    # Organize classes by semester
    organized_classes, current_semester = organize_classes_by_semester(all_classes)
    
    # Get classes for current semester (all year levels)
    current_semester_classes = []
    if current_semester in organized_classes:
        for year_level, classes in organized_classes[current_semester].items():
            current_semester_classes.extend(classes)
    
    # Sort current semester classes by year level then name
    current_semester_classes.sort(key=lambda x: (x.year or 1, x.class_name))
    
    # Serialize current semester classes for JavaScript
    current_semester_classes_json = [serialize_class(cls) for cls in current_semester_classes]
    
    # Statistics for the dashboard
    total_classes = len(all_classes)
    active_classes = len([c for c in all_classes if c.is_active])
    current_semester_count = len(current_semester_classes)
    
    # Get year level breakdown for current semester
    year_level_counts = defaultdict(int)
    if current_semester in organized_classes:
        for year in organized_classes[current_semester].keys():
            year_level_counts[year] = len(organized_classes[current_semester][year])
    
    # Also serialize all organized classes for JavaScript
    organized_classes_json = {}
    for semester, year_levels in organized_classes.items():
        organized_classes_json[semester] = {}
        for year, classes in year_levels.items():
            organized_classes_json[semester][year] = [serialize_class(cls) for cls in classes]
    
    return render_template(
        'lecturer/reports.html',
        organized_classes=organized_classes,
        organized_classes_json=organized_classes_json,  # Add serialized version
        current_semester_classes=current_semester_classes,
        current_semester_classes_json=current_semester_classes_json,  # Add serialized version
        current_semester=current_semester,
        year_level_counts=dict(year_level_counts),
        total_classes=total_classes,
        active_classes=active_classes,
        current_semester_count=current_semester_count,
        page_title='Reports & Analytics'
    )


@reports_bp.route('/api/classes/filter')
@login_required
@active_account_required
def filter_classes():
    """
    API endpoint to filter classes by semester (all year levels)
    """
    semester = request.args.get('semester')
    year_level = request.args.get('year_level', type=int)  # Optional filter by year level
    
    all_classes = current_user.get_assigned_classes()
    
    # Filter classes
    filtered_classes = []
    for cls in all_classes:
        # Extract semester from format: year.semester (e.g., '3.1' -> semester '1')
        if cls.semester and '.' in cls.semester:
            parts = cls.semester.split('.')
            cls_semester = parts[1] if len(parts) > 1 else None
        elif cls.semester:
            cls_semester = cls.semester
        else:
            cls_semester = None
        
        # Only include valid semesters
        if cls_semester not in ['1', '2']:
            continue
            
        if semester and cls_semester != semester:
            continue
        if year_level and cls.year != year_level:
            continue
            
        filtered_classes.append({
            'class_id': cls.class_id,
            'class_name': cls.class_name,
            'course_code': cls.course_code,
            'year': cls.year,
            'semester': cls.semester,
            'display_name': f"{cls.class_id} - {cls.class_name} (Year {cls.year}, {cls.course_code})"
        })
    
    # Sort by year level then class name
    filtered_classes.sort(key=lambda x: (x['year'] or 0, x['class_name']))
    
    return jsonify({
        'classes': filtered_classes,
        'count': len(filtered_classes)
    })


# ========== Session Reports ==========

@reports_bp.route('/session/<int:session_id>')
@login_required
@active_account_required
@owns_session
def session_report(session_id):
    """Display session report in browser"""
    report_data = report_service.generate_session_summary(
        session_id, 
        current_user.instructor_id
    )
    
    if not report_data:
        flash('Session not found or access denied', 'error')
        return redirect(url_for('reports.index'))
    
    return render_template(
        'lecturer/report_session.html',
        report_data=report_data,
        page_title='Session Report'
    )


@reports_bp.route('/session/<int:session_id>/export/<format>')
@login_required
@active_account_required
@owns_session
def export_session_report(session_id, format):
    """Export session report in specified format"""
    report_data = report_service.generate_session_summary(
        session_id,
        current_user.instructor_id
    )
    
    if not report_data:
        flash('Session not found or access denied', 'error')
        return redirect(url_for('reports.index'))
    
    session_info = report_data['session_info']
    filename_base = f"session_report_{session_id}_{session_info['date']}"
    
    if format == 'pdf':
        pdf_content = pdf_generator.generate_session_report(report_data)
        return send_file(
            io.BytesIO(pdf_content),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{filename_base}.pdf"
        )
    
    elif format == 'excel':
        excel_content = excel_generator.generate_session_report(report_data)
        return send_file(
            io.BytesIO(excel_content),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{filename_base}.xlsx"
        )
    
    elif format == 'csv':
        csv_content = _generate_session_csv(report_data)
        return send_file(
            io.BytesIO(csv_content.encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{filename_base}.csv"
        )
    
    elif format == 'json':
        return jsonify(report_data)
    
    else:
        flash('Invalid export format', 'error')
        return redirect(url_for('reports.session_report', session_id=session_id))


# ========== Class Reports ==========

@reports_bp.route('/class/<class_id>')
@login_required
@active_account_required
@owns_class
def class_report(class_id):
    """Display class summary report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not end_date:
        end_date = datetime.now().date().strftime('%Y-%m-%d')
    if not start_date:
        start_date = (datetime.now().date() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    report_data = report_service.generate_class_summary(
        class_id,
        current_user.instructor_id,
        start_date,
        end_date
    )
    
    if not report_data:
        flash('Class not found or access denied', 'error')
        return redirect(url_for('reports.index'))
    
    return render_template(
        'lecturer/report_class.html',
        report_data=report_data,
        start_date=start_date,
        end_date=end_date,
        page_title='Class Report'
    )


@reports_bp.route('/class/<class_id>/export/<format>')
@login_required
@active_account_required
@owns_class
def export_class_report(class_id, format):
    """Export class report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    report_data = report_service.generate_class_summary(
        class_id,
        current_user.instructor_id,
        start_date,
        end_date
    )
    
    if not report_data:
        flash('Class not found or access denied', 'error')
        return redirect(url_for('reports.index'))
    
    filename_base = f"class_report_{class_id}_{datetime.now().strftime('%Y%m%d')}"
    
    if format == 'pdf':
        pdf_content = pdf_generator.generate_class_report(report_data)
        return send_file(
            io.BytesIO(pdf_content),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{filename_base}.pdf"
        )
    
    elif format == 'excel':
        excel_content = excel_generator.generate_class_report(report_data)
        return send_file(
            io.BytesIO(excel_content),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{filename_base}.xlsx"
        )
    
    elif format == 'csv':
        csv_content = _generate_class_csv(report_data)
        return send_file(
            io.BytesIO(csv_content.encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{filename_base}.csv"
        )
    
    elif format == 'json':
        return jsonify(report_data)
    
    else:
        flash('Invalid export format', 'error')
        return redirect(url_for('reports.class_report', class_id=class_id))


# ========== Student Reports ==========

@reports_bp.route('/student/<student_id>/class/<class_id>')
@login_required
@active_account_required
@owns_class
def student_report(student_id, class_id):
    """Display individual student report"""
    report_data = report_service.generate_student_report(
        student_id,
        class_id,
        current_user.instructor_id
    )
    
    if not report_data:
        flash('Student or class not found, or access denied', 'error')
        return redirect(url_for('reports.index'))
    
    return render_template(
        'lecturer/report_student.html',
        report_data=report_data,
        page_title='Student Report'
    )


@reports_bp.route('/student/<student_id>/class/<class_id>/export/<format>')
@login_required
@active_account_required
@owns_class
def export_student_report(student_id, class_id, format):
    """Export student report"""
    report_data = report_service.generate_student_report(
        student_id,
        class_id,
        current_user.instructor_id
    )
    
    if not report_data:
        flash('Student or class not found, or access denied', 'error')
        return redirect(url_for('reports.index'))
    
    filename_base = f"student_report_{student_id}_{datetime.now().strftime('%Y%m%d')}"
    
    if format == 'pdf':
        pdf_content = pdf_generator.generate_student_report(report_data)
        return send_file(
            io.BytesIO(pdf_content),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{filename_base}.pdf"
        )
    
    elif format == 'excel':
        excel_content = excel_generator.generate_student_report(report_data)
        return send_file(
            io.BytesIO(excel_content),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{filename_base}.xlsx"
        )
    
    elif format == 'csv':
        csv_content = _generate_student_csv(report_data)
        return send_file(
            io.BytesIO(csv_content.encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{filename_base}.csv"
        )
    
    elif format == 'json':
        return jsonify(report_data)
    
    else:
        flash('Invalid export format', 'error')
        return redirect(url_for('reports.student_report', 
                               student_id=student_id, 
                               class_id=class_id))


# ========== Trend Analysis ==========

@reports_bp.route('/trend/<class_id>')
@login_required
@active_account_required
@owns_class
def trend_analysis(class_id):
    """Display attendance trend analysis"""
    days = request.args.get('days', 30, type=int)
    
    report_data = report_service.generate_trend_analysis(
        class_id,
        current_user.instructor_id,
        days
    )
    
    if not report_data:
        flash('Class not found or access denied', 'error')
        return redirect(url_for('reports.index'))
    
    return render_template(
        'lecturer/report_trend.html',
        report_data=report_data,
        days=days,
        page_title='Trend Analysis'
    )


# ========== Alert Reports ==========

@reports_bp.route('/alerts')
@login_required
@active_account_required
def low_attendance_alerts():
    """Display low attendance alerts"""
    threshold = request.args.get('threshold', 75, type=int)
    
    report_data = report_service.generate_low_attendance_alert(
        current_user.instructor_id,
        threshold
    )
    
    if not report_data:
        flash('Unable to generate alert report', 'error')
        return redirect(url_for('reports.index'))
    
    return render_template(
        'lecturer/report_alerts.html',
        report_data=report_data,
        threshold=threshold,
        page_title='Low Attendance Alerts'
    )


@reports_bp.route('/alerts/export/<format>')
@login_required
@active_account_required
def export_alerts(format):
    """Export low attendance alerts"""
    threshold = request.args.get('threshold', 75, type=int)
    
    report_data = report_service.generate_low_attendance_alert(
        current_user.instructor_id,
        threshold
    )
    
    if not report_data:
        flash('Unable to generate alert report', 'error')
        return redirect(url_for('reports.index'))
    
    filename_base = f"low_attendance_alerts_{datetime.now().strftime('%Y%m%d')}"
    
    if format == 'pdf':
        pdf_content = pdf_generator.generate_alert_report(report_data)
        return send_file(
            io.BytesIO(pdf_content),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{filename_base}.pdf"
        )
    
    elif format == 'excel':
        excel_content = excel_generator.generate_alert_report(report_data)
        return send_file(
            io.BytesIO(excel_content),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{filename_base}.xlsx"
        )
    
    elif format == 'csv':
        csv_content = _generate_alert_csv(report_data)
        return send_file(
            io.BytesIO(csv_content.encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{filename_base}.csv"
        )
    
    elif format == 'json':
        return jsonify(report_data)
    
    else:
        flash('Invalid export format', 'error')
        return redirect(url_for('reports.low_attendance_alerts'))


# ========== Bulk Export ==========

@reports_bp.route('/bulk-export')
@login_required
@active_account_required
def bulk_export():
    """Display bulk export options"""
    classes = current_user.get_assigned_classes()
    
    return render_template(
        'lecturer/bulk_export.html',
        classes=classes,
        page_title='Bulk Export'
    )


@reports_bp.route('/bulk-export/process', methods=['POST'])
@login_required
@active_account_required
def process_bulk_export():
    """Process bulk export request"""
    export_type = request.form.get('export_type')
    export_format = request.form.get('format')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    
    if export_type == 'all_classes':
        classes = current_user.get_assigned_classes()
        class_ids = [c.class_id for c in classes]
    else:
        class_ids = request.form.getlist('class_ids')
    
    if not class_ids:
        flash('Please select at least one class', 'error')
        return redirect(url_for('reports.bulk_export'))
    
    if export_format == 'excel':
        excel_content = _generate_bulk_excel(
            class_ids,
            current_user.instructor_id,
            start_date,
            end_date
        )
        
        return send_file(
            io.BytesIO(excel_content),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"bulk_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
    
    else:
        flash('Only Excel format is supported for bulk export', 'error')
        return redirect(url_for('reports.bulk_export'))


# ========== API Endpoints ==========

@reports_bp.route('/api/class/<class_id>/chart-data')
@login_required
@active_account_required
@owns_class
def get_class_chart_data(class_id):
    """Get chart data for class report"""
    days = request.args.get('days', 30, type=int)
    
    trend_data = report_service.generate_trend_analysis(
        class_id,
        current_user.instructor_id,
        days
    )
    
    if not trend_data:
        return jsonify({'error': 'Data not available'}), 404
    
    return jsonify({
        'labels': [d['date'] for d in trend_data['trend']['data']],
        'data': [d['attendance_rate'] for d in trend_data['trend']['data']],
        'trend_direction': trend_data['trend']['direction']
    })


@reports_bp.route('/api/sessions/recent')
@login_required
@active_account_required
def get_recent_session():
    """Get most recent session for quick report"""
    from app.models import ClassSession
    from app.models.class_model import Class
    
    session = ClassSession.query.join(Class).filter(
        Class.class_id.in_([c.class_id for c in current_user.get_assigned_classes()]),
        ClassSession.status.in_(['completed', 'ongoing'])
    ).order_by(ClassSession.date.desc(), ClassSession.start_time.desc()).first()
    
    if session:
        return jsonify({
            'session_id': session.session_id,
            'class_name': session.class_.class_name,
            'date': session.date.isoformat() if session.date else None,
            'start_time': session.start_time.strftime('%H:%M') if session.start_time else None
        })
    
    return jsonify({'error': 'No recent sessions'}), 404


@reports_bp.route('/api/classes/<class_id>/sessions')
@login_required
@active_account_required
@owns_class
def get_class_sessions(class_id):
    """Get all sessions for a class"""
    from app.models import ClassSession
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = ClassSession.query.filter_by(class_id=class_id)
    
    if start_date:
        query = query.filter(ClassSession.date >= start_date)
    if end_date:
        query = query.filter(ClassSession.date <= end_date)
    
    sessions = query.order_by(ClassSession.date.desc()).all()
    
    return jsonify([{
        'session_id': s.session_id,
        'date': s.date.isoformat() if s.date else None,
        'start_time': s.start_time.strftime('%H:%M') if s.start_time else None,
        'end_time': s.end_time.strftime('%H:%M') if s.end_time else None,
        'status': s.status,
        'attendance_count': s.attendance_count,
        'total_students': s.total_students
    } for s in sessions])


@reports_bp.route('/api/classes/<class_id>/students')
@login_required
@active_account_required
@owns_class
def get_class_students(class_id):
    """Get all students in a class"""
    from app.models.class_model import Class
    
    class_obj = Class.query.get(class_id)
    if not class_obj:
        return jsonify({'error': 'Class not found'}), 404
    
    return jsonify([{
        'student_id': s.student_id,
        'fname': s.fname,
        'lname': s.lname,
        'email': s.email
    } for s in class_obj.students])


@reports_bp.route('/api/session/<int:session_id>/stats')
@login_required
@active_account_required
@owns_session
def get_session_stats(session_id):
    """Get statistics for a session"""
    report_data = report_service.generate_session_summary(
        session_id,
        current_user.instructor_id
    )
    
    if not report_data:
        return jsonify({'error': 'Session not found'}), 404
    
    return jsonify(report_data['statistics'])


# ========== Helper Functions ==========

def _generate_session_csv(report_data):
    """Generate CSV for session report"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Student ID', 'Name', 'Email', 'Status', 'Timestamp', 'Method', 'Confidence'])
    
    for student in report_data['students']:
        writer.writerow([
            student['student_id'],
            student['name'],
            student['email'],
            student['status'],
            student['timestamp'],
            student['method'],
            student['confidence']
        ])
    
    return output.getvalue()


def _generate_class_csv(report_data):
    """Generate CSV for class report"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Student ID', 'Name', 'Email', 'Present', 'Late', 'Absent', 'Attendance Rate'])
    
    for student in report_data['students']:
        writer.writerow([
            student['student_id'],
            student['name'],
            student['email'],
            student['present'],
            student['late'],
            student['absent'],
            f"{student['attendance_rate']:.1f}%"
        ])
    
    return output.getvalue()


def _generate_student_csv(report_data):
    """Generate CSV for student report"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Session ID', 'Date', 'Start Time', 'End Time', 'Status', 'Timestamp', 'Method'])
    
    for record in report_data['attendance_history']:
        writer.writerow([
            record['session_id'],
            record['date'],
            record['start_time'],
            record['end_time'],
            record['status'],
            record['timestamp'],
            record['method']
        ])
    
    return output.getvalue()


def _generate_alert_csv(report_data):
    """Generate CSV for alert report"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Student ID', 'Name', 'Email', 'Class', 'Course', 
                    'Attended', 'Total Sessions', 'Attendance Rate'])
    
    for student in report_data['students']:
        writer.writerow([
            student['student_id'],
            student['name'],
            student['email'],
            student['class_name'],
            student['course_code'],
            student['attended'],
            student['total_sessions'],
            f"{student['attendance_rate']:.1f}%"
        ])
    
    return output.getvalue()


def _generate_bulk_excel(class_ids, instructor_id, start_date, end_date):
    """Generate Excel file with multiple class reports"""
    from openpyxl import Workbook
    
    wb = Workbook()
    wb.remove(wb.active)
    
    for class_id in class_ids:
        report_data = report_service.generate_class_summary(
            class_id,
            instructor_id,
            start_date,
            end_date
        )
        
        if report_data:
            class_name = report_data['class_info']['class_name']
            sheet_name = class_name[:31].replace('/', '-').replace('\\', '-')
            
            ws = wb.create_sheet(sheet_name)
            
            ws['A1'] = "Class Report"
            ws['A2'] = f"Class: {class_name}"
            ws['A3'] = f"Period: {start_date} to {end_date}"
            
            ws['A5'] = "Student ID"
            ws['B5'] = "Name"
            ws['C5'] = "Present"
            ws['D5'] = "Late"
            ws['E5'] = "Absent"
            ws['F5'] = "Rate"
            
            row = 6
            for student in report_data['students']:
                ws[f'A{row}'] = student['student_id']
                ws[f'B{row}'] = student['name']
                ws[f'C{row}'] = student['present']
                ws[f'D{row}'] = student['late']
                ws[f'E{row}'] = student['absent']
                ws[f'F{row}'] = f"{student['attendance_rate']:.1f}%"
                row += 1
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()