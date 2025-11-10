"""
Excel Generator - Generate professional Excel reports with formatting
Supports multiple sheets, charts, and conditional formatting
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


class ExcelGenerator:
    """Generate Excel reports with professional formatting"""
    
    def __init__(self):
        # Define color scheme
        self.colors = {
            'header': 'FF2C3E50',
            'success': 'FF27AE60',
            'warning': 'FFF39C12',
            'danger': 'FFE74C3C',
            'info': 'FF3498DB',
            'light_gray': 'FFECF0F1',
            'white': 'FFFFFFFF'
        }
        
        # Define styles
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFFFF')
        self.title_font = Font(name='Arial', size=16, bold=True, color='FF2C3E50')
        self.normal_font = Font(name='Arial', size=10)
        
        self.header_fill = PatternFill(start_color=self.colors['header'], 
                                       end_color=self.colors['header'], 
                                       fill_type='solid')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
    
    def generate_session_report(self, report_data, output_path=None):
        """
        Generate Excel report for a single session
        
        Args:
            report_data: Session report data from ReportService
            output_path: Path to save Excel file (optional)
            
        Returns:
            bytes: Excel content if output_path is None, else None
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Overview sheet
        self._create_session_overview_sheet(wb, report_data)
        
        # Create Student Details sheet
        self._create_student_details_sheet(wb, report_data)
        
        # Create Statistics sheet with chart
        self._create_statistics_sheet(wb, report_data)
        
        return self._save_workbook(wb, output_path)
    
    def generate_class_report(self, report_data, output_path=None):
        """Generate Excel report for class summary"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create Overview sheet
        self._create_class_overview_sheet(wb, report_data)
        
        # Create Sessions sheet
        self._create_sessions_sheet(wb, report_data)
        
        # Create Student Performance sheet
        self._create_student_performance_sheet(wb, report_data)
        
        # Create Trend Analysis sheet
        self._create_trend_sheet(wb, report_data)
        
        return self._save_workbook(wb, output_path)
    
    def generate_student_report(self, report_data, output_path=None):
        """Generate Excel report for individual student"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create Student Info sheet
        self._create_student_info_sheet(wb, report_data)
        
        # Create Attendance History sheet
        self._create_attendance_history_sheet(wb, report_data)
        
        # Create Statistics sheet
        self._create_student_statistics_sheet(wb, report_data)
        
        return self._save_workbook(wb, output_path)
    
    def generate_alert_report(self, report_data, output_path=None):
        """Generate Excel report for low attendance alerts"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create Alert Summary sheet
        self._create_alert_summary_sheet(wb, report_data)
        
        # Create At-Risk Students sheet
        self._create_at_risk_students_sheet(wb, report_data)
        
        return self._save_workbook(wb, output_path)
    
    # ========== Session Report Sheets ==========
    
    def _create_session_overview_sheet(self, wb, data):
        """Create overview sheet for session report"""
        ws = wb.create_sheet("Overview", 0)
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "Session Attendance Report"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_alignment
        
        session_info = data['session_info']
        stats = data['statistics']
        
        # Session Information
        ws['A3'] = "Session Information"
        ws['A3'].font = Font(size=14, bold=True)
        
        info_data = [
            ['Session ID:', session_info['session_id']],
            ['Class:', session_info['class_name']],
            ['Course:', f"{session_info['course_code']} - {session_info['course_name']}"],
            ['Date:', session_info['date']],
            ['Time:', f"{session_info['start_time']} - {session_info['end_time']}"],
            ['Status:', session_info['status'].upper()]
        ]
        
        row = 4
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Statistics Summary
        ws['D3'] = "Statistics"
        ws['D3'].font = Font(size=14, bold=True)
        
        stats_data = [
            ['Total Expected:', stats['total_expected']],
            ['Present:', stats['present']],
            ['Late:', stats['late']],
            ['Absent:', stats['absent']],
            ['Excused:', stats['excused']],
            ['Attendance Rate:', f"{stats['attendance_rate']}%"]
        ]
        
        row = 4
        for label, value in stats_data:
            ws[f'D{row}'] = label
            ws[f'D{row}'].font = Font(bold=True)
            ws[f'E{row}'] = value
            
            # Color code the values
            if label == 'Present:':
                ws[f'E{row}'].fill = PatternFill(start_color='FFD4EDDA', 
                                                 end_color='FFD4EDDA', 
                                                 fill_type='solid')
            elif label == 'Absent:':
                ws[f'E{row}'].fill = PatternFill(start_color='FFF8D7DA', 
                                                 end_color='FFF8D7DA', 
                                                 fill_type='solid')
            elif label == 'Late:':
                ws[f'E{row}'].fill = PatternFill(start_color='FFFFF3CD', 
                                                 end_color='FFFFF3CD', 
                                                 fill_type='solid')
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
    
    def _create_student_details_sheet(self, wb, data):
        """Create student details sheet"""
        ws = wb.create_sheet("Student Details")
        
        # Headers
        headers = ['Student ID', 'Name', 'Email', 'Status', 'Time', 'Method', 'Confidence', 'Notes']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Data
        students = data['students']
        for row, student in enumerate(students, start=2):
            ws.cell(row=row, column=1).value = student['student_id']
            ws.cell(row=row, column=2).value = student['name']
            ws.cell(row=row, column=3).value = student['email']
            ws.cell(row=row, column=4).value = student['status']
            ws.cell(row=row, column=5).value = student['timestamp']
            ws.cell(row=row, column=6).value = student['method']
            ws.cell(row=row, column=7).value = student['confidence']
            ws.cell(row=row, column=8).value = student['notes']
            
            # Apply status color
            status_cell = ws.cell(row=row, column=4)
            if student['status'] == 'Present':
                status_cell.fill = PatternFill(start_color='FFD4EDDA', 
                                              end_color='FFD4EDDA', 
                                              fill_type='solid')
            elif student['status'] == 'Late':
                status_cell.fill = PatternFill(start_color='FFFFF3CD', 
                                              end_color='FFFFF3CD', 
                                              fill_type='solid')
            elif student['status'] == 'Absent':
                status_cell.fill = PatternFill(start_color='FFF8D7DA', 
                                              end_color='FFF8D7DA', 
                                              fill_type='solid')
            
            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.thin_border
        
        # Auto-adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        ws.column_dimensions['B'].width = 25  # Name column
        ws.column_dimensions['C'].width = 30  # Email column
        ws.column_dimensions['H'].width = 40  # Notes column
    
    def _create_statistics_sheet(self, wb, data):
        """Create statistics sheet with pie chart"""
        ws = wb.create_sheet("Statistics")
        
        # Data for chart
        ws['A1'] = "Status"
        ws['B1'] = "Count"
        
        stats = data['statistics']
        ws['A2'] = "Present"
        ws['B2'] = stats['present']
        ws['A3'] = "Late"
        ws['B3'] = stats['late']
        ws['A4'] = "Absent"
        ws['B4'] = stats['absent']
        
        # Style headers
        ws['A1'].font = self.header_font
        ws['B1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws['B1'].fill = self.header_fill
        
        # Create pie chart
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=4)
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=4)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Attendance Distribution"
        
        ws.add_chart(pie, "D2")
    
    # ========== Class Report Sheets ==========
    
    def _create_class_overview_sheet(self, wb, data):
        """Create overview sheet for class report"""
        ws = wb.create_sheet("Overview", 0)
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "Class Attendance Report"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_alignment
        
        class_info = data['class_info']
        period = data['period']
        stats = data['statistics']
        
        # Class Information
        ws['A3'] = "Class Information"
        ws['A3'].font = Font(size=14, bold=True)
        
        info_data = [
            ['Class ID:', class_info['class_id']],
            ['Class Name:', class_info['class_name']],
            ['Course:', f"{class_info['course_code']} - {class_info['course_name']}"],
            ['Year:', class_info['year']],
            ['Semester:', class_info['semester']],
            ['Total Students:', class_info['total_students']]
        ]
        
        row = 4
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Period Information
        ws['D3'] = "Report Period"
        ws['D3'].font = Font(size=14, bold=True)
        
        period_data = [
            ['Start Date:', period['start_date'] or 'All Time'],
            ['End Date:', period['end_date'] or 'All Time'],
            ['Total Sessions:', period['total_sessions']],
            ['Completed Sessions:', period['completed_sessions']]
        ]
        
        row = 4
        for label, value in period_data:
            ws[f'D{row}'] = label
            ws[f'D{row}'].font = Font(bold=True)
            ws[f'E{row}'] = value
            row += 1
        
        # Overall Statistics
        ws['A12'] = "Overall Statistics"
        ws['A12'].font = Font(size=14, bold=True)
        
        stats_data = [
            ['Total Present:', stats['total_present']],
            ['Total Late:', stats['total_late']],
            ['Total Absent:', stats['total_absent']],
            ['Average Attendance Rate:', f"{stats['average_attendance_rate']}%"]
        ]
        
        row = 13
        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
    
    def _create_sessions_sheet(self, wb, data):
        """Create sessions list sheet"""
        ws = wb.create_sheet("Sessions")
        
        # Headers
        headers = ['Session ID', 'Date', 'Start Time', 'End Time', 'Status', 
                   'Present', 'Total', 'Attendance Rate']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Data
        sessions = data['sessions']
        for row, session in enumerate(sessions, start=2):
            ws.cell(row=row, column=1).value = session['session_id']
            ws.cell(row=row, column=2).value = session['date']
            ws.cell(row=row, column=3).value = session['start_time']
            ws.cell(row=row, column=4).value = session['end_time']
            ws.cell(row=row, column=5).value = session['status']
            ws.cell(row=row, column=6).value = session['attendance_count']
            ws.cell(row=row, column=7).value = session['total_students']
            ws.cell(row=row, column=8).value = f"{session['attendance_rate']:.1f}%"
            
            # Color code attendance rate
            rate_cell = ws.cell(row=row, column=8)
            if session['attendance_rate'] >= 80:
                rate_cell.fill = PatternFill(start_color='FFD4EDDA', 
                                            end_color='FFD4EDDA', 
                                            fill_type='solid')
            elif session['attendance_rate'] >= 60:
                rate_cell.fill = PatternFill(start_color='FFFFF3CD', 
                                            end_color='FFFFF3CD', 
                                            fill_type='solid')
            else:
                rate_cell.fill = PatternFill(start_color='FFF8D7DA', 
                                            end_color='FFF8D7DA', 
                                            fill_type='solid')
            
            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.thin_border
        
        # Auto-adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_student_performance_sheet(self, wb, data):
        """Create student performance sheet"""
        ws = wb.create_sheet("Student Performance")
        
        # Headers
        headers = ['Student ID', 'Name', 'Email', 'Present', 'Late', 
                   'Absent', 'Attendance Rate']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Data
        students = data['students']
        for row, student in enumerate(students, start=2):
            ws.cell(row=row, column=1).value = student['student_id']
            ws.cell(row=row, column=2).value = student['name']
            ws.cell(row=row, column=3).value = student['email']
            ws.cell(row=row, column=4).value = student['present']
            ws.cell(row=row, column=5).value = student['late']
            ws.cell(row=row, column=6).value = student['absent']
            ws.cell(row=row, column=7).value = f"{student['attendance_rate']:.1f}%"
            
            # Color code attendance rate
            rate_cell = ws.cell(row=row, column=7)
            if student['attendance_rate'] >= 75:
                rate_cell.fill = PatternFill(start_color='FFD4EDDA', 
                                            end_color='FFD4EDDA', 
                                            fill_type='solid')
            elif student['attendance_rate'] >= 60:
                rate_cell.fill = PatternFill(start_color='FFFFF3CD', 
                                            end_color='FFFFF3CD', 
                                            fill_type='solid')
            else:
                rate_cell.fill = PatternFill(start_color='FFF8D7DA', 
                                            end_color='FFF8D7DA', 
                                            fill_type='solid')
            
            # Apply borders
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.thin_border
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        for col in ['D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15
    
    def _create_trend_sheet(self, wb, data):
        """Create trend analysis sheet with line chart"""
        ws = wb.create_sheet("Trend Analysis")
        
        # Headers
        ws['A1'] = "Date"
        ws['B1'] = "Attendance Rate"
        
        ws['A1'].font = self.header_font
        ws['B1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws['B1'].fill = self.header_fill
        
        # Data
        sessions = data['sessions'][-20:]  # Last 20 sessions
        for row, session in enumerate(sessions, start=2):
            ws.cell(row=row, column=1).value = session['date']
            ws.cell(row=row, column=2).value = session['attendance_rate']
        
        # Create line chart
        chart = LineChart()
        chart.title = "Attendance Trend"
        chart.y_axis.title = "Attendance Rate (%)"
        chart.x_axis.title = "Date"
        
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(sessions)+1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(sessions)+1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D2")
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
    
    # ========== Student Report Sheets ==========
    
    def _create_student_info_sheet(self, wb, data):
        """Create student information sheet"""
        ws = wb.create_sheet("Student Info", 0)
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Student Attendance Report"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_alignment
        
        student_info = data['student_info']
        class_info = data['class_info']
        stats = data['statistics']
        
        # Student Information
        ws['A3'] = "Student Information"
        ws['A3'].font = Font(size=14, bold=True)
        
        info_data = [
            ['Student ID:', student_info['student_id']],
            ['Name:', student_info['name']],
            ['Email:', student_info['email']],
            ['Phone:', student_info['phone']],
            ['Course:', student_info['course']],
            ['Year:', student_info['year']],
            ['Semester:', student_info['semester']]
        ]
        
        row = 4
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Class Information
        ws['A13'] = "Class Information"
        ws['A13'].font = Font(size=14, bold=True)
        
        ws['A14'] = "Class:"
        ws['A14'].font = Font(bold=True)
        ws['B14'] = class_info['class_name']
        
        ws['A15'] = "Course:"
        ws['A15'].font = Font(bold=True)
        ws['B15'] = f"{class_info['course_code']} - {class_info['course_name']}"
        
        # Statistics
        ws['A18'] = "Attendance Statistics"
        ws['A18'].font = Font(size=14, bold=True)
        
        stats_data = [
            ['Total Sessions:', stats['total_sessions']],
            ['Present:', stats['present']],
            ['Late:', stats['late']],
            ['Absent:', stats['absent']],
            ['Excused:', stats['excused']],
            ['Attendance Rate:', f"{stats['attendance_rate']}%"]
        ]
        
        row = 19
        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            
            # Highlight attendance rate
            if label == 'Attendance Rate:':
                cell = ws[f'B{row}']
                if stats['attendance_rate'] >= 75:
                    cell.fill = PatternFill(start_color='FFD4EDDA', 
                                          end_color='FFD4EDDA', 
                                          fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFF8D7DA', 
                                          end_color='FFF8D7DA', 
                                          fill_type='solid')
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_attendance_history_sheet(self, wb, data):
        """Create attendance history sheet"""
        ws = wb.create_sheet("Attendance History")
        
        # Headers
        headers = ['Session ID', 'Date', 'Start Time', 'End Time', 
                   'Status', 'Marked At', 'Method', 'Notes']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Data
        history = data['attendance_history']
        for row, record in enumerate(history, start=2):
            ws.cell(row=row, column=1).value = record['session_id']
            ws.cell(row=row, column=2).value = record['date']
            ws.cell(row=row, column=3).value = record['start_time']
            ws.cell(row=row, column=4).value = record['end_time']
            ws.cell(row=row, column=5).value = record['status']
            ws.cell(row=row, column=6).value = record['timestamp']
            ws.cell(row=row, column=7).value = record['method']
            ws.cell(row=row, column=8).value = record['notes']
            
            # Apply status color
            status_cell = ws.cell(row=row, column=5)
            if record['status'] == 'Present':
                status_cell.fill = PatternFill(start_color='FFD4EDDA', 
                                              end_color='FFD4EDDA', 
                                              fill_type='solid')
            elif record['status'] == 'Late':
                status_cell.fill = PatternFill(start_color='FFFFF3CD', 
                                              end_color='FFFFF3CD', 
                                              fill_type='solid')
            elif record['status'] == 'Absent':
                status_cell.fill = PatternFill(start_color='FFF8D7DA', 
                                              end_color='FFF8D7DA', 
                                              fill_type='solid')
            
            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.thin_border
        
        # Auto-adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        ws.column_dimensions['H'].width = 40  # Notes column
    
    def _create_student_statistics_sheet(self, wb, data):
        """Create student statistics sheet with chart"""
        ws = wb.create_sheet("Statistics")
        
        stats = data['statistics']
        
        # Data for chart
        ws['A1'] = "Status"
        ws['B1'] = "Count"
        
        ws['A2'] = "Present"
        ws['B2'] = stats['present']
        ws['A3'] = "Late"
        ws['B3'] = stats['late']
        ws['A4'] = "Absent"
        ws['B4'] = stats['absent']
        
        # Style headers
        ws['A1'].font = self.header_font
        ws['B1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws['B1'].fill = self.header_fill
        
        # Create pie chart
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=4)
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=4)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Attendance Distribution"
        
        ws.add_chart(pie, "D2")
    
    # ========== Alert Report Sheets ==========
    
    def _create_alert_summary_sheet(self, wb, data):
        """Create alert summary sheet"""
        ws = wb.create_sheet("Alert Summary", 0)
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "⚠️ Low Attendance Alert Report"
        ws['A1'].font = Font(size=16, bold=True, color='FFE74C3C')
        ws['A1'].alignment = self.center_alignment
        
        alert_info = data['alert_info']
        
        # Alert Information
        ws['A3'] = "Alert Details"
        ws['A3'].font = Font(size=14, bold=True)
        
        ws['A4'] = "Threshold:"
        ws['A4'].font = Font(bold=True)
        ws['B4'] = f"{alert_info['threshold']}%"
        
        ws['A5'] = "Students At Risk:"
        ws['A5'].font = Font(bold=True)
        ws['B5'] = alert_info['total_at_risk']
        ws['B5'].fill = PatternFill(start_color='FFF8D7DA', 
                                    end_color='FFF8D7DA', 
                                    fill_type='solid')
        
        ws['A6'] = "Instructor:"
        ws['A6'].font = Font(bold=True)
        ws['B6'] = alert_info['instructor_name']
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_at_risk_students_sheet(self, wb, data):
        """Create at-risk students list sheet"""
        ws = wb.create_sheet("At-Risk Students")
        
        # Headers
        headers = ['Student ID', 'Name', 'Email', 'Class', 'Course', 
                   'Attended', 'Missed', 'Attendance Rate']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Data
        students = data['students']
        for row, student in enumerate(students, start=2):
            ws.cell(row=row, column=1).value = student['student_id']
            ws.cell(row=row, column=2).value = student['name']
            ws.cell(row=row, column=3).value = student['email']
            ws.cell(row=row, column=4).value = student['class_name']
            ws.cell(row=row, column=5).value = student['course_code']
            ws.cell(row=row, column=6).value = f"{student['attended']}/{student['total_sessions']}"
            ws.cell(row=row, column=7).value = student['sessions_missed']
            ws.cell(row=row, column=8).value = f"{student['attendance_rate']:.1f}%"
            
            # Color code based on severity
            rate_cell = ws.cell(row=row, column=8)
            if student['attendance_rate'] < 50:
                rate_cell.fill = PatternFill(start_color='FFE74C3C', 
                                            end_color='FFE74C3C', 
                                            fill_type='solid')
                rate_cell.font = Font(color='FFFFFFFF', bold=True)
            elif student['attendance_rate'] < 65:
                rate_cell.fill = PatternFill(start_color='FFF8D7DA', 
                                            end_color='FFF8D7DA', 
                                            fill_type='solid')
            else:
                rate_cell.fill = PatternFill(start_color='FFFFF3CD', 
                                            end_color='FFFFF3CD', 
                                            fill_type='solid')
            
            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.thin_border
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
    
    # ========== Helper Methods ==========
    
    def _save_workbook(self, wb, output_path=None):
        """Save workbook to file or return as bytes"""
        if output_path:
            wb.save(output_path)
            return None
        else:
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()